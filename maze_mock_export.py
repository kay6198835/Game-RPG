import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── helpers ──────────────────────────────────────────────────────────────────

def header_style(ws, row, cols, text, bg="1F4E79"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22

def col_header(ws, row, col, text, bg="2E75B6"):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 18

def data_cell(ws, row, col, value, bg=None, bold=False, align="center", color="000000"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=color, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    thin = Side(style="thin", color="BFBFBF")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

STATUS_COLOR = {
    "DISABLE":  "D9D9D9",  # gray
    "ENEBLE":   "FF0000",  # red  (carved OUT)
    "BE_OPEN":  "FFFFFF",  # white (carved IN)
    "OPEN":     "00B050",  # green
    "CLOSE":    "595959",  # dark gray
}

STATUS_TEXT_COLOR = {
    "DISABLE":  "808080",
    "ENEBLE":   "FFFFFF",
    "BE_OPEN":  "000000",
    "OPEN":     "FFFFFF",
    "CLOSE":    "FFFFFF",
}

def status_cell(ws, row, col, status):
    bg   = STATUS_COLOR.get(status, "FFFFFF")
    fc   = STATUS_TEXT_COLOR.get(status, "000000")
    cell = ws.cell(row=row, column=col, value=status)
    cell.font = Font(bold=True, color=fc, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="BFBFBF")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── SHEET 1: Cell Data ────────────────────────────────────────────────────────

ws1 = wb.active
ws1.title = "1_Cell_Data"

cells = [
    # idx, row, col, TOP,      BOTTOM,   LEFT,     RIGHT
    (0,   0, 0, "DISABLE", "ENEBLE",  "DISABLE","DISABLE"),
    (1,   0, 1, "DISABLE", "BE_OPEN", "DISABLE","ENEBLE"),
    (2,   0, 2, "DISABLE", "ENEBLE",  "BE_OPEN","DISABLE"),
    (3,   1, 0, "BE_OPEN", "DISABLE", "DISABLE","ENEBLE"),
    (4,   1, 1, "ENEBLE",  "DISABLE", "BE_OPEN","DISABLE"),
    (5,   1, 2, "BE_OPEN", "ENEBLE",  "DISABLE","DISABLE"),
    (6,   2, 0, "DISABLE", "DISABLE", "DISABLE","BE_OPEN"),
    (7,   2, 1, "DISABLE", "DISABLE", "ENEBLE", "BE_OPEN"),
    (8,   2, 2, "BE_OPEN", "DISABLE", "ENEBLE", "DISABLE"),
]

header_style(ws1, 1, 8, "PHASE 1 — MazeGenerator.Generator(3, 3)  →  Cell[] Gird")

col_headers = ["Index", "row", "col", "TOP", "BOTTOM", "LEFT", "RIGHT", "Visited"]
for ci, h in enumerate(col_headers, 1):
    col_header(ws1, 2, ci, h)

for r, (idx, row, col, top, bot, left, right) in enumerate(cells, 3):
    data_cell(ws1, r, 1, idx,   bold=True, bg="EBF3FB")
    data_cell(ws1, r, 2, row)
    data_cell(ws1, r, 3, col)
    status_cell(ws1, r, 4, top)
    status_cell(ws1, r, 5, bot)
    status_cell(ws1, r, 6, left)
    status_cell(ws1, r, 7, right)
    data_cell(ws1, r, 8, "true", bg="E2EFDA")

# legend
ws1.cell(row=14, column=1, value="Legend:")
ws1.cell(row=14, column=1).font = Font(bold=True)
legend = [
    ("ENEBLE",  STATUS_COLOR["ENEBLE"],  STATUS_TEXT_COLOR["ENEBLE"],  "Cell chủ động đục tường (màu đỏ trên minimap)"),
    ("BE_OPEN", STATUS_COLOR["BE_OPEN"], STATUS_TEXT_COLOR["BE_OPEN"], "Cell bên kia đục sang (màu trắng trên minimap)"),
    ("DISABLE", STATUS_COLOR["DISABLE"], STATUS_TEXT_COLOR["DISABLE"], "Tường kín — không spawn door"),
]
for li, (name, bg, fc, desc) in enumerate(legend, 15):
    c = ws1.cell(row=li, column=2, value=name)
    c.font = Font(bold=True, color=fc)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center")
    thin = Side(style="thin", color="999999")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws1.cell(row=li, column=3, value=desc).font = Font(italic=True, size=9)

set_col_widths(ws1, [8, 6, 6, 12, 12, 12, 12, 10])

# ── SHEET 2: Maze Visual ──────────────────────────────────────────────────────

ws2 = wb.create_sheet("2_Maze_Visual")
header_style(ws2, 1, 10, "Maze Layout — 3×3 Grid  (view từ trên xuống)")

# column labels
for ci, label in enumerate(["", "col=0", "col=1", "col=2"], 1):
    col_header(ws2, 2, ci, label, bg="2E75B6" if ci > 1 else "1F4E79")

# draw maze cells
maze = [
    # (label, top_open, right_open, bottom_open, left_open, bg)
    ["[0,0]\nRoom[0]\nStart", False, False, True,  False, "FFF2CC"],  # row0
    ["[0,1]\nRoom[1]",        False, True,  False, False, "DEEAF1"],
    ["[0,2]\nRoom[2]",        False, False, True,  True,  "DEEAF1"],
    ["[1,0]\nRoom[3]",        True,  True,  False, False, "DEEAF1"],  # row1
    ["[1,1]\nRoom[4]",        True,  False, False, True,  "DEEAF1"],
    ["[1,2]\nRoom[5]",        True,  False, True,  False, "DEEAF1"],
    ["[2,0]\nRoom[6]",        False, True,  False, False, "F4CCCC"],  # row2
    ["[2,1]\nRoom[7]",        False, False, False, True,  "F4CCCC"],
    ["[2,2]\nRoom[8]",        True,  False, False, True,  "F4CCCC"],
]

cell_rows = [[maze[0], maze[1], maze[2]],
             [maze[3], maze[4], maze[5]],
             [maze[6], maze[7], maze[8]]]

for ri, mrow in enumerate(cell_rows):
    row_label = ws2.cell(row=3+ri, column=1, value=f"row={ri}")
    row_label.font = Font(bold=True, color="FFFFFF", size=10)
    row_label.fill = PatternFill("solid", fgColor="2E75B6")
    row_label.alignment = Alignment(horizontal="center", vertical="center")

    for ci, (label, top, right, bottom, left, bg) in enumerate(mrow):
        doors = []
        if top:    doors.append("↑TOP")
        if right:  doors.append("RIGHT→")
        if bottom: doors.append("↓BOT")
        if left:   doors.append("←LEFT")
        text = label + "\n" + "  ".join(doors)
        cell = ws2.cell(row=3+ri, column=2+ci, value=text)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thick = Side(style="medium", color="000000")
        thin  = Side(style="thin",   color="999999")
        cell.border = Border(
            top=thick    if not top    else Side(style="dashed", color="FF0000"),
            right=thick  if not right  else Side(style="dashed", color="FF0000"),
            bottom=thick if not bottom else Side(style="dashed", color="FF0000"),
            left=thick   if not left   else Side(style="dashed", color="FF0000"),
        )
        ws2.row_dimensions[3+ri].height = 55

set_col_widths(ws2, [10, 22, 22, 22])

# DFS path note
ws2.cell(row=7, column=1, value="DFS Path:").font = Font(bold=True)
path_text = "(0,0)→↓(1,0)→→(1,1)→↑(0,1)→→(0,2)→↓(1,2)→↓(2,2)→←(2,1)→←(2,0)"
ws2.cell(row=8, column=1, value=path_text).font = Font(italic=True, size=10, color="1F4E79")
ws2.merge_cells(start_row=8, start_column=1, end_row=8, end_column=5)

# ── SHEET 3: RoomGrid (RoomCell) ──────────────────────────────────────────────

ws3 = wb.create_sheet("3_RoomGrid_WorldPos")
header_style(ws3, 1, 9, "RoomGridController — RoomCell World Positions & Doors  (GAME_SCALE=3, LENGTH_ROOM=10)")

col_hs = ["Index","row","col","World X","World Y","Scale","TOP door","BOTTOM door","LEFT door","RIGHT door"]
for ci, h in enumerate(col_hs, 1):
    col_header(ws3, 2, ci, h)

room_data = [
    # idx row col wx   wy    top         bot        left       right
    (0,  0, 0,  0,   0,   "DISABLE", "ENEBLE",  "DISABLE", "DISABLE"),
    (1,  0, 1,  30,  0,   "DISABLE", "BE_OPEN", "DISABLE", "ENEBLE"),
    (2,  0, 2,  60,  0,   "DISABLE", "ENEBLE",  "BE_OPEN", "DISABLE"),
    (3,  1, 0,  0,  -30,  "BE_OPEN", "DISABLE", "DISABLE", "ENEBLE"),
    (4,  1, 1,  30, -30,  "ENEBLE",  "DISABLE", "BE_OPEN", "DISABLE"),
    (5,  1, 2,  60, -30,  "BE_OPEN", "ENEBLE",  "DISABLE", "DISABLE"),
    (6,  2, 0,  0,  -60,  "DISABLE", "DISABLE", "DISABLE", "BE_OPEN"),
    (7,  2, 1,  30, -60,  "DISABLE", "DISABLE", "ENEBLE",  "BE_OPEN"),
    (8,  2, 2,  60, -60,  "BE_OPEN", "DISABLE", "ENEBLE",  "DISABLE"),
]

for r, (idx, row, col, wx, wy, top, bot, left, right) in enumerate(room_data, 3):
    bg_row = "EBF3FB" if r % 2 == 0 else "FFFFFF"
    data_cell(ws3, r, 1,  idx,  bold=True, bg="D6E4F0")
    data_cell(ws3, r, 2,  row,  bg=bg_row)
    data_cell(ws3, r, 3,  col,  bg=bg_row)
    data_cell(ws3, r, 4,  wx,   bg="FFF2CC", bold=True)
    data_cell(ws3, r, 5,  wy,   bg="FFF2CC", bold=True)
    data_cell(ws3, r, 6,  3,    bg=bg_row)
    status_cell(ws3, r, 7,  top)
    status_cell(ws3, r, 8,  bot)
    status_cell(ws3, r, 9,  left)
    status_cell(ws3, r, 10, right)

# formula note
note_row = len(room_data) + 4
ws3.cell(row=note_row, column=1, value="Formula:").font = Font(bold=True)
ws3.cell(row=note_row+1, column=1, value="World X = col * GAME_SCALE * LENGTH_ROOM = col * 30").font = Font(italic=True, size=9)
ws3.cell(row=note_row+2, column=1, value="World Y = -row * GAME_SCALE * LENGTH_ROOM = -row * 30").font = Font(italic=True, size=9)
ws3.merge_cells(start_row=note_row+1, start_column=1, end_row=note_row+1, end_column=5)
ws3.merge_cells(start_row=note_row+2, start_column=1, end_row=note_row+2, end_column=5)

set_col_widths(ws3, [8, 6, 6, 10, 10, 8, 12, 12, 12, 12])

# ── SHEET 4: MapGrid (MapCell) ────────────────────────────────────────────────

ws4 = wb.create_sheet("4_MapGrid_MiniMap")
header_style(ws4, 1, 8, "MapGridController — MapCell MiniMap Positions  (spacing = GAME_SCALE * LENGTH_CELL * 2 = 6)")

col_hs = ["Index","row","col","Pos X (from origin)","Pos Y (from origin)","_top active","_bottom active","_left active","_right active"]
for ci, h in enumerate(col_hs, 1):
    col_header(ws4, 2, ci, h)

map_data = [
    # idx row col  dx  dy    top    bot    left   right
    (0,  0, 0,  0,   0,   False, True,  False, False),
    (1,  0, 1,  6,   0,   False, False, False, True),
    (2,  0, 2,  12,  0,   False, True,  False, False),
    (3,  1, 0,  0,  -6,   False, False, False, True),
    (4,  1, 1,  6,  -6,   True,  False, False, False),
    (5,  1, 2,  12, -6,   False, True,  False, False),
    (6,  2, 0,  0,  -12,  False, False, False, False),
    (7,  2, 1,  6,  -12,  False, False, True,  False),
    (8,  2, 2,  12, -12,  False, False, True,  False),
]

YES_BG = "00B050"
NO_BG  = "D9D9D9"

for r, (idx, row, col, dx, dy, top, bot, left, right) in enumerate(map_data, 3):
    bg_row = "EBF3FB" if r % 2 == 0 else "FFFFFF"
    data_cell(ws4, r, 1, idx,  bold=True, bg="D6E4F0")
    data_cell(ws4, r, 2, row,  bg=bg_row)
    data_cell(ws4, r, 3, col,  bg=bg_row)
    data_cell(ws4, r, 4, f"Ox + {dx}", bg="FFF2CC")
    data_cell(ws4, r, 5, f"Oy + ({dy})", bg="FFF2CC")
    for ci, flag in enumerate([top, bot, left, right], 6):
        val = "SetActive(true)" if flag else "—"
        bg  = "C6EFCE" if flag else "F2F2F2"
        fc  = "276221" if flag else "808080"
        data_cell(ws4, r, ci, val, bg=bg, color=fc)

ws4.cell(row=14, column=1, value="Note:").font = Font(bold=True)
n = ws4.cell(row=15, column=1, value="MapCell chỉ SetActive(true) child indicator khi door status == ENEBLE (cell chủ động).\nBE_OPEN side không hiển thị indicator trên minimap.")
n.font = Font(italic=True, size=9)
n.alignment = Alignment(wrap_text=True)
ws4.row_dimensions[15].height = 30
ws4.merge_cells(start_row=15, start_column=1, end_row=15, end_column=6)

set_col_widths(ws4, [8, 6, 6, 16, 16, 18, 18, 18, 18])

# ── SHEET 5: Navigation Example ───────────────────────────────────────────────

ws5 = wb.create_sheet("5_Navigation_Example")
header_style(ws5, 1, 6, "Navigation Example — Player đi BOTTOM từ Room[0] → Room[3]")

steps = [
    ("Event fired",         "EventManager.Emit(ON_PLAYER_ON_DOOR, Vector2.down)"),
    ("direction",           "Vector2.down = (0, -1)"),
    ("BaseGrid.GetNext()",  "direction.y flip: (0,-1) → (0,+1)"),
    ("positionNextRoom",    "_current.GetGridPosition() + (0,+1) = (0,0)+(0,1) = (0,1)"),
    ("index",               "1 * Columns + 0 = 1*3+0 = 3"),
    ("_next",               "RoomCell[3]  at world (0, -30)"),
    ("direction flip back", "(0,+1) → (0,-1)  [restore]"),
    ("OnAfterGetNext()",    "current=Room[0], next=Room[3], dir=(0,-1)"),
    ("UpdateStatusDoor",    "Room[0].GetDoor(BOTTOM).OpenDoor()  → status=OPEN"),
    ("GetStartDoorPosition","called with -direction = -(0,-1) = (0,+1) = TOP"),
    ("GetDoor(TOP)",        "Room[3] has TOP=BE_OPEN  → door found ✓"),
    ("nextDoor.OpenDoor()", "Room[3].TOP door → status=OPEN"),
    ("offset calc",         "direction*(PADDING=3)*(SCALE=3)*1.1 = (0,+1)*9.9 = (0,9.9)"),
    ("StartDoorPosition",   "topDoor.worldPos − (0, 9.9)  ≈  (0, 3.6)  [~9.9u inside room]"),
    ("Player teleport",     "player.transform.position = (0, 3.6)"),
    ("MiniMap",             "ON_PLAYER_ON_DOOR → CellMapController.GetNext(down) → MapCell[3]"),
    ("Avatar",              "Avatar.position = MapCell[3].transform.position  = (Ox, Oy-6)"),
    ("ON_LOAD_MAP",         "EventManager.Emit(ON_LOAD_MAP) → _current = MapCell[3]"),
]

col_header(ws5, 2, 1, "Step", bg="1F4E79")
col_header(ws5, 2, 2, "Detail", bg="1F4E79")

for ri, (step, detail) in enumerate(steps, 3):
    bg = "EBF3FB" if ri % 2 == 0 else "FFFFFF"
    if "teleport" in step.lower() or "StartDoor" in step:
        bg = "E2EFDA"
    if "Event" in step or "ON_" in detail:
        bg = "FFF2CC"
    data_cell(ws5, ri, 1, step,   bold=True, bg=bg, align="left")
    data_cell(ws5, ri, 2, detail, bg=bg, align="left")

set_col_widths(ws5, [28, 65])

# ── SHEET 6: STATUS_DOOR Reference ───────────────────────────────────────────

ws6 = wb.create_sheet("6_StatusDoor_Reference")
header_style(ws6, 1, 5, "STATUS_DOOR Enum Reference")

col_header(ws6, 2, 1, "Value")
col_header(ws6, 2, 2, "Name")
col_header(ws6, 2, 3, "Dùng ở đâu")
col_header(ws6, 2, 4, "RoomCell hiển thị")
col_header(ws6, 2, 5, "MapCell indicator")

ref = [
    (0, "DISABLE",  "Cell mặc định — không có cửa",         "Không spawn DoorController",    "Không hiện"),
    (1, "ENEBLE",   "Cell chủ động đục tường (carve out)",   "Spawn door, màu đỏ",            "SetActive(true)"),
    (2, "BE_OPEN",  "Cell bị đục từ bên kia (carve in)",    "Spawn door, màu trắng",         "Không hiện"),
    (3, "OPEN",     "Runtime: cửa mở, player đi qua được",  "Trigger ON, visual open",       "N/A"),
    (4, "CLOSE",    "Runtime: cửa đóng, locked",             "Trigger OFF, visual locked",    "N/A"),
]
for ri, (val, name, where, room, mapc) in enumerate(ref, 3):
    status_cell(ws6, ri, 1, str(val))
    status_cell(ws6, ri, 2, name)
    data_cell(ws6, ri, 3, where, align="left")
    data_cell(ws6, ri, 4, room,  align="left")
    data_cell(ws6, ri, 5, mapc,  align="left")

set_col_widths(ws6, [8, 12, 40, 32, 20])

# ── Save ─────────────────────────────────────────────────────────────────────

out = "/home/user/Game-RPG/maze_mock_data.xlsx"
wb.save(out)
print(f"Saved: {out}")
